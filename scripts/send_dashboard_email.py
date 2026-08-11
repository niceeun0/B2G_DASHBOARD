#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
send_dashboard_email.py
=========================================================
build_dashboard.py가 재생성한 'outputs/정부지원사업_대시보드.xlsx'를
이메일로 첨부해서 발송합니다. bizinfo_bot.py의 send_email()(Gmail SMTP,
재시도 로직 포함)을 그대로 재사용합니다 - 같은 MAIL_USER/EMAIL_PASS/
MAIL_RECEIVER 환경변수(시크릿)를 사용합니다.

[첨부파일 경량화]
원본 파일에는 5개 시트가 들어있는데, 그중 '공공기관현황_원본'과
'1년치지원사업_원본'은 절대 안 바뀌는 고정 데이터라 매일 통째로 다시
보낼 필요가 없습니다 (필요하면 저장소의 원본 파일에서 확인 가능).
이 두 시트를 뺀 '경량판'을 만들어서 zip으로 압축한 뒤 보내면, 원본
대비 약 60% 작아집니다 (단순 zip 압축만으로는 7%밖에 안 줄어듦 -
xlsx 자체가 이미 내부적으로 압축된 형식이라서). 외부 회사 메일
게이트웨이가 큰 첨부파일을 막는 경우를 줄이기 위한 조치입니다.

워크플로우에서 build_dashboard.py 다음 단계로 실행하면 됩니다:
  python scripts/build_dashboard.py
  python scripts/send_dashboard_email.py
=========================================================
"""

import os
import io
import sys
import zipfile
from datetime import datetime, timedelta, timezone

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import bizinfo_bot as bot  # send_email() 재사용
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(os.path.dirname(BASE_DIR), "outputs")
DASHBOARD_PATH = os.path.join(OUT_DIR, "정부지원사업_대시보드.xlsx")
STATUS_PATH = os.path.join(os.path.dirname(BASE_DIR), "data", "last_run_status.json")

# 메일에는 이 시트들만 포함 (매일 갱신되는/실제로 매일 확인이 필요한 시트).
# '공공기관현황_원본'과 '1년치지원사업_원본'은 고정 데이터라 제외합니다.
EMAIL_SHEETS = ["대시보드_우선순위_랭킹", "공공기관별_컨택디렉토리", "매일업로드_원천데이터"]


def read_status():
    """bizinfo_bot.py가 기록한 실행 상태를 읽습니다. 못 읽으면 빈 dict."""
    if not os.path.exists(STATUS_PATH):
        return {}
    try:
        import json
        with open(STATUS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        bot.log(f"실행 상태 파일을 읽지 못했습니다(항상 대시보드를 보냅니다): {e}")
        return {}


def mark_dashboard_email_sent(status):
    """이 스크립트가 오늘 메일을 보냈다는 걸 상태 파일에 남깁니다 - 하루 중
    재시도 스케줄로 다시 실행되더라도 같은 내용을 중복 발송하지 않기 위함."""
    try:
        import json
        status["dashboard_email_sent"] = True
        with open(STATUS_PATH, "w", encoding="utf-8") as f:
            json.dump(status, f, ensure_ascii=False)
    except Exception as e:
        bot.log(f"실행 상태 파일 갱신 실패(치명적이지 않음): {e}")


def build_lightweight_xlsx_bytes(path):
    """고정 원본 시트를 뺀 경량판 xlsx를 만들어 바이트로 반환."""
    wb = openpyxl.load_workbook(path)
    light = openpyxl.Workbook()
    light.remove(light.active)

    for sheet_name in EMAIL_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        src = wb[sheet_name]
        dst = light.create_sheet(sheet_name)
        for row in src.iter_rows():
            for cell in row:
                dst.cell(row=cell.row, column=cell.column, value=bot.sanitize_for_excel(cell.value))
        # 열 너비도 최대한 유지 (가독성)
        for col_letter, dim in src.column_dimensions.items():
            if dim.width:
                dst.column_dimensions[col_letter].width = dim.width

    buf = io.BytesIO()
    light.save(buf)
    return buf.getvalue()


def zip_bytes(data, inner_filename):
    """바이트를 zip으로 압축해서 반환."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        zf.writestr(inner_filename, data)
    return buf.getvalue()


def build_summary_html(today_str):
    """대시보드 상위 요약 정보를 메일 본문에 짧게 넣습니다 (첨부파일 전체 열지 않아도
    바로 감이 오도록)."""
    try:
        import pandas as pd
        df = pd.read_excel(DASHBOARD_PATH, sheet_name="대시보드_우선순위_랭킹", header=3)
        total = len(df)
        contactable = int((df["담당자_이메일_통합모음"].fillna("").str.strip() != "").sum())
        sent = int(df["발송여부"].fillna("").str.upper().eq("Y").sum())
        top5 = df.sort_values("우선순위_종합점수", ascending=False).head(5)
        rows_html = "".join(
            f"<tr><td style='padding:6px 10px;border-bottom:1px solid #e5e7eb;'>{r['제안순위']}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #e5e7eb;'>{r['지점명']}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #e5e7eb;'>{r['우선순위_종합점수']}</td></tr>"
            for _, r in top5.iterrows()
        )
        summary = f"""
        <div style="margin:14px 0;font-size:13.5px;color:#374151;">
          전체 {total}개 기관 · 연락 가능 {contactable}개 · 발송 완료 {sent}개
        </div>
        <table style="width:100%;border-collapse:collapse;margin-bottom:14px;">
          <thead><tr style="background:#111827;color:#fff;">
            <th style="padding:6px 10px;text-align:left;">순위</th>
            <th style="padding:6px 10px;text-align:left;">기관</th>
            <th style="padding:6px 10px;text-align:left;">점수</th>
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
        """
        return summary
    except Exception as e:
        bot.log(f"요약 정보 생성 실패(첨부파일은 정상 발송됨): {e}")
        return ""


def send_no_new_notice(today_str, new_count):
    """신규 등록이 0건인 날(주말/공휴일 다음날 등) 짧은 알림만 보냅니다."""
    subject = f"[정부지원사업 대시보드] {today_str} - 신규 등록 없음"
    html_body = f"""
    <html><body style="margin:0;padding:0;background:#f3f4f6;font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif;">
      <div style="max-width:640px;margin:0 auto;padding:24px 16px;">
        <h2 style="color:#111827;margin:0 0 6px 0;">📭 오늘은 신규 등록 공고가 없습니다</h2>
        <div style="color:#6b7280;font-size:13px;margin-bottom:16px;">{today_str} 기준</div>
        <div style="text-align:center;padding:16px;background:#f9fafb;border-radius:8px;
                    font-size:13.5px;color:#374151;">
          기업마당에 신규 등록된 지원사업 공고가 0건이라(주말/공휴일 다음날에
          흔히 있는 상황입니다), 대시보드 전체 파일은 오늘 다시 보내지 않습니다.
          누적 데이터는 변동이 없으니 최근에 받으신 대시보드 파일을 그대로
          참고하시면 됩니다.
        </div>
      </div>
    </body></html>
    """
    bot.send_email(subject, html_body)


def main():
    if not os.path.exists(DASHBOARD_PATH):
        bot.log(f"[치명적 오류] 대시보드 파일이 없습니다: {DASHBOARD_PATH}")
        sys.exit(1)

    kst = timezone(timedelta(hours=9))
    today_str = datetime.now(kst).strftime("%Y-%m-%d")

    status = read_status()
    new_count = status.get("new_count")

    # 하루 중 재시도 스케줄(예: 08시 실패 -> 14시 재시도)로 다시 실행된
    # 경우, 오늘 이미 대시보드 메일을 보냈다면 중복 발송하지 않고 건너뜁니다.
    if status.get("dashboard_email_sent"):
        bot.log(
            f"오늘({status.get('target_date')}) 대시보드 메일은 이미 발송됐습니다 "
            "(하루 중 재시도 스케줄로 재실행된 것으로 보임) - 중복 발송 방지를 위해 건너뜁니다."
        )
        return

    if new_count == 0:
        bot.log(f"오늘 신규 등록 0건 확인 - 짧은 알림만 보내고 대시보드 첨부는 건너뜁니다.")
        try:
            send_no_new_notice(today_str, new_count)
            mark_dashboard_email_sent(status)
        except Exception as e:
            bot.log(f"[치명적 오류] 알림 메일 발송 실패: {e}")
            sys.exit(1)
        return

    xlsx_filename = f"정부지원사업_대시보드_{today_str}.xlsx"
    zip_filename = f"정부지원사업_대시보드_{today_str}.zip"
    subject = f"[정부지원사업 대시보드] {today_str} 갱신본"

    full_size = os.path.getsize(DASHBOARD_PATH)
    light_bytes = build_lightweight_xlsx_bytes(DASHBOARD_PATH)
    zip_bytes_data = zip_bytes(light_bytes, xlsx_filename)
    bot.log(
        f"첨부파일 경량화: 원본 {full_size:,} bytes -> 경량판+zip {len(zip_bytes_data):,} bytes "
        f"({(1 - len(zip_bytes_data) / full_size) * 100:.0f}% 감소, "
        f"제외된 시트: 공공기관현황_원본/1년치지원사업_원본 - 저장소 원본 파일에서 확인 가능)"
    )

    summary_html = build_summary_html(today_str)
    html_body = f"""
    <html><body style="margin:0;padding:0;background:#f3f4f6;font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif;">
      <div style="max-width:640px;margin:0 auto;padding:24px 16px;">
        <h2 style="color:#111827;margin:0 0 6px 0;">📊 정부지원사업 대시보드 갱신본</h2>
        <div style="color:#6b7280;font-size:13px;margin-bottom:16px;">{today_str} 기준</div>
        {summary_html}
        <div style="text-align:center;padding:14px;background:#eef2ff;border-radius:8px;
                    font-size:13.5px;color:#3730a3;">
          📎 첨부된 <b>{zip_filename}</b> 압축파일을 풀면 엑셀 파일이 나옵니다.<br>
          (파일 용량을 줄이기 위해, 매일 안 바뀌는 '공공기관현황_원본'/'1년치지원사업_원본'
          시트는 이 메일에서는 빼고 보내드립니다. 필요하시면 저장소의 원본 파일을 확인해주세요.)
        </div>
      </div>
    </body></html>
    """

    try:
        bot.send_email(
            subject, html_body,
            attachment_bytes=zip_bytes_data,
            attachment_filename=zip_filename,
        )
        mark_dashboard_email_sent(status)
    except Exception as e:
        bot.log(f"[치명적 오류] 대시보드 메일 발송 실패: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
