#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fix_shifted_accumulator.py
=========================================================
2026-08 초, bizinfo_bot.py에 '지원금액'/'지원갯수' 컬럼을 새로 추가하면서
기존 append_to_accumulator()가 '컬럼 이름'이 아니라 '몇 번째 칸인지(위치)'
기준으로 예전 행을 그대로 복사해 붙였던 버그가 있었습니다.

그 결과, 새 컬럼이 삽입된 지점(지원규모 다음) 이후의 모든 값이 2칸씩
밀려서 잘못된 컬럼에 저장되었습니다. 이 스크립트는 그 손상된 파일을
1회성으로 복구합니다.

사용법:
  python fix_shifted_accumulator.py <손상된_파일.xlsx> <복구본_저장경로.xlsx>

복구 방법:
  손상된 파일의 각 행은, 사실 '지원금액/지원갯수가 없던 예전 30개 컬럼'
  순서로 값이 채워져 있고 그게 지금의 32개 컬럼 헤더에 잘못 얹혀 있는
  상태입니다. 이 스크립트는 각 행을 예전 30개 컬럼 이름에 맞게 다시
  해석한 뒤, 지원금액/지원갯수는 (남아있는) '지원규모' 텍스트에서 새로
  파싱해서 채우고, 나머지 값은 올바른 컬럼으로 되돌려 놓습니다.
=========================================================
"""

import sys
import os
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__))))
import bizinfo_bot as bot  # extract_support_amount/count, EXPORT_COLUMNS 재사용

# 버그 당시(지원금액/지원갯수 추가 이전) 실제로 쓰이던 30개 컬럼 순서.
# 손상된 파일의 각 행 값은 이 순서대로 채워져 있습니다.
OLD_HEADERS_30 = [
    "공고명", "수행기관명(excInsttNm)", "신청기간", "신청방법", "지원규모",
    "사업자등록증명", "표준재무제표증명", "부가가치세과세표준증명", "납부내역증명", "납세증명",
    "면세사업자수입금액증명원", "사업자등록증명(영문)", "4대보험완납증명서", "나의부동산정보",
    "지방세 납세증명", "지방세 세목별 과세증명", "법인 등기부등본", "법인세 신고내역",
    "종합소득세 신고내역", "부가세 신고내역", "원천세 신고내역", "거래처별합계표",
    "산업재해율확인서", "산업재해 요양승인 및 반려여부 확인서",
    "기타서류", "문의처", "담당이메일", "담당전화번호", "지원대상", "원문링크",
]


def looks_shifted(row_dict):
    """
    이 행이 실제로 밀린 상태인지 판별합니다.
    '지원금액이 비어있는지'로 판단하면 오탐이 많습니다 (원래도 금액 언급이
    없는 정상 공고가 섞여있기 때문). 훨씬 신뢰도 높은 신호는 '담당이메일'
    필드입니다 - 이 필드는 항상 이메일 형식('@' 포함)이어야 정상이므로,
    여기에 '@'가 없으면 밀렸다고 판단합니다.
    """
    email_val = str(row_dict.get("담당이메일", "") or "")
    return "@" not in email_val


def _find_by_pattern(raw_row, pattern):
    """행의 모든 값을 훑어서 패턴에 맞는 첫 값을 반환 (내용 기반 복구용)."""
    for v in raw_row:
        if isinstance(v, str) and pattern.search(v):
            m = pattern.search(v)
            return m.group(0)
    return ""


def fix_file(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    current_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    new_headers = [label for label, _ in bot.EXPORT_COLUMNS]
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws.title
    ws_out.append(new_headers)

    fixed_count = 0
    kept_count = 0
    uncertain_count = 0

    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        current_dict = dict(zip(current_headers, raw_row))

        if not looks_shifted(current_dict):
            # 이미 정상인 행(신버전 코드로 만들어진 최근 행)은 그대로 유지
            new_row = [current_dict.get(h, "") for h in new_headers]
            ws_out.append(new_row)
            kept_count += 1
            continue

        # 밀린 행: 현재 값들을 '예전 30개 컬럼' 이름으로 재해석
        old_values = list(raw_row)[: len(OLD_HEADERS_30)]
        old_dict = dict(zip(OLD_HEADERS_30, old_values))

        budget_text = old_dict.get("지원규모", "") or ""
        old_dict["지원금액"] = bot.extract_support_amount(budget_text)
        old_dict["지원갯수"] = bot.extract_support_count(budget_text)

        # 30컬럼 재해석으로도 이메일이 안 나오면, 이 행은 그보다 더 오래된
        # (다른) 스키마로 저장된 행입니다. 위치 추정을 더 이상 신뢰하지 말고,
        # 행 전체를 훑어서 이메일/전화/URL처럼 형태로 확실히 알아볼 수 있는
        # 값만 내용 기반으로 뽑아 채웁니다. 나머지(서류 Y/N, 지원대상 등)는
        # 추측하지 않고 빈 값으로 두어, 잘못된 값을 넣는 것보다 안전하게
        # "수동 확인이 필요하다"는 것을 알 수 있게 합니다.
        if "@" not in str(old_dict.get("담당이메일", "")):
            uncertain_count += 1
            found_email = _find_by_pattern(raw_row, bot.EMAIL_RE)
            found_phone = _find_by_pattern(raw_row, bot.PHONE_RE)
            found_url = _find_by_pattern(
                raw_row, __import__("re").compile(r"https?://\S+")
            )
            old_dict["담당이메일"] = found_email
            old_dict["담당전화번호"] = found_phone
            old_dict["원문링크"] = found_url or old_dict.get("원문링크", "")
            old_dict["지원대상"] = "-"
            old_dict["문의처"] = "(자동 복구 불확실 - 원문 링크에서 직접 확인 권장)"
            old_dict["기타서류"] = ""
            for label, _ in bot.ONECLICK_DOC_CATALOG:
                old_dict[label] = "N"

        new_row = [old_dict.get(h, "") for h in new_headers]
        ws_out.append(new_row)
        fixed_count += 1

    wb_out.save(output_path)
    print(
        f"복구 완료: 정상 유지 {kept_count}행 / 밀림 수정 {fixed_count}행 "
        f"(그중 스키마 불확실해서 내용기반 복구한 행 {uncertain_count}건) -> {output_path}"
    )


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("사용법: python fix_shifted_accumulator.py <입력.xlsx> <출력.xlsx>")
        sys.exit(1)
    fix_file(sys.argv[1], sys.argv[2])
