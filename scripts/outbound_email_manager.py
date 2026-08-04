#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
outbound_email_manager.py
=========================================================
'정부지원사업_대시보드.xlsx'의 우선순위 랭킹을 기준으로,
아직 제안 메일을 보내지 않은 담당자만 골라내는 스크립트입니다.

핵심 아이디어 (엑셀 체크박스 방식):
- 대시보드_우선순위_랭킹 시트에 '발송여부'/'발송일자' 컬럼이 있습니다.
  엑셀을 열어서 '발송여부' 셀을 Y로 직접 체크해도 되고(드롭다운 제공),
  이 스크립트의 mark-sent 명령으로 자동 기록해도 됩니다. 둘 다 같은
  컬럼을 보고/씁니다.
- build_dashboard.py가 대시보드를 재생성할 때도 이 값은 유지됩니다
  (지점명 기준으로 기존 값을 읽어와서 새 대시보드에 다시 넣어줌).
- outputs/발송이력.csv에는 '누가 언제 어떤 캠페인으로' 보냈는지 상세
  이력이 별도로 계속 쌓입니다 (감사/백업용 - 대시보드의 발송여부는
  '현재 상태'만 보여주고, 이 CSV가 '전체 히스토리'를 담습니다).
- 실제 메일 발송 자체는 이 스크립트가 하지 않습니다. 발송 대상 목록만
  뽑아주고, 사용자가 메일머지/직접발송 등으로 보낸 뒤 'mark-sent'로
  기록하는 구조입니다.

사용법:
  # 1) 다음 발송 대상 50건 뽑기 (우선순위 높은 순, 발송여부가 Y가 아닌 곳만)
  python outbound_email_manager.py next --top 50

  # 2) 실제로 메일을 보낸 뒤, 대시보드의 발송여부/발송일자를 자동 기록
  python outbound_email_manager.py mark-sent \
      --from outputs/발송대상_다음배치.xlsx --campaign "2026-08 1차 제안"

  # 3) 특정 기관 몇 개만 수동으로 발송완료 처리하고 싶을 때
  python outbound_email_manager.py mark-sent --branches "대구테크노파크,제주테크노파크" --campaign "수동추가"

  # 4) 지금까지 발송 현황 요약 보기
  python outbound_email_manager.py status
=========================================================
"""

import os
import argparse
from datetime import datetime

import pandas as pd
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(os.path.dirname(BASE_DIR), "outputs")
DASHBOARD_PATH = os.path.join(OUT_DIR, "정부지원사업_대시보드.xlsx")
HISTORY_PATH = os.path.join(OUT_DIR, "발송이력.csv")

RANK_SHEET = "대시보드_우선순위_랭킹"
HEADER_ROW_1INDEXED = 4  # 실제 헤더가 4번째 줄(제목/부제/빈줄 다음)
HEADER_ROW_0INDEXED = 3  # pandas.read_excel(header=...)용


def log(msg):
    print(f"[outbound] {msg}", flush=True)


# =========================================================
# 발송이력 CSV (감사 로그용 - 대시보드 컬럼과 별개로 계속 쌓음)
# =========================================================
def load_history(path=HISTORY_PATH):
    if not os.path.exists(path):
        return pd.DataFrame(columns=["기관명", "지점명", "이메일", "발송일자", "캠페인명"])
    return pd.read_csv(path, encoding="utf-8-sig")


def save_history(df, path=HISTORY_PATH):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


# =========================================================
# 대시보드 로드/기록
# =========================================================
def load_rank_sheet(dashboard_path=DASHBOARD_PATH):
    return pd.read_excel(dashboard_path, sheet_name=RANK_SHEET, header=HEADER_ROW_0INDEXED)


def is_sent(value):
    return isinstance(value, str) and value.strip().upper() == "Y"


def write_send_status(branch_names, campaign, dashboard_path=DASHBOARD_PATH):
    """
    지정된 지점명들의 '발송여부'를 Y로, '발송일자'를 오늘 날짜로 대시보드
    엑셀 파일에 직접 기록합니다 (엑셀에서 수동으로 체크한 것과 동일한 효과).
    """
    today = datetime.now().strftime("%Y-%m-%d")
    wb = openpyxl.load_workbook(dashboard_path)
    ws = wb[RANK_SHEET]

    header_cells = list(ws.iter_rows(min_row=HEADER_ROW_1INDEXED, max_row=HEADER_ROW_1INDEXED))[0]
    headers = [c.value for c in header_cells]
    try:
        branch_col = headers.index("지점명") + 1
        status_col = headers.index("발송여부") + 1
        date_col = headers.index("발송일자") + 1
    except ValueError as e:
        log(f"대시보드에서 필요한 컬럼을 찾지 못했습니다: {e}")
        return 0

    branch_set = set(branch_names)
    updated = 0
    for row_idx in range(HEADER_ROW_1INDEXED + 1, ws.max_row + 1):
        branch_val = ws.cell(row=row_idx, column=branch_col).value
        if branch_val in branch_set:
            ws.cell(row=row_idx, column=status_col).value = "Y"
            ws.cell(row=row_idx, column=date_col).value = today
            updated += 1

    wb.save(dashboard_path)
    log(f"대시보드의 발송여부/발송일자 {updated}개 기관에 기록 완료")
    return updated


# =========================================================
# 명령: next - 다음 발송 대상 뽑기
# =========================================================
def cmd_next(args):
    df = load_rank_sheet(args.dashboard)
    log(f"대시보드 {len(df)}개 기관 로드")

    results = []
    for _, row in df.iterrows():
        if is_sent(row.get("발송여부")):
            continue  # 이미 Y로 체크된(발송완료) 기관은 제외
        emails_str = row.get("담당자_이메일_통합모음", "") or ""
        if not isinstance(emails_str, str) or not emails_str.strip():
            continue
        results.append({
            "제안순위": row.get("제안순위"),
            "기관명": row.get("기관명"),
            "지점명": row.get("지점명"),
            "상위기관_주무부처_지자체": row.get("상위기관_주무부처_지자체"),
            "우선순위_종합점수": row.get("우선순위_종합점수"),
            "발송대상이메일": emails_str,
            "발송대상이메일_개수": len([e for e in emails_str.split(",") if e.strip()]),
            "대표전화": row.get("대표전화"),
        })

    if not results:
        log("발송 대상이 없습니다 (모두 발송 완료로 체크되어 있거나, 이메일이 확보된 기관이 없습니다).")
        return

    out_df = pd.DataFrame(results).sort_values("우선순위_종합점수", ascending=False)
    if args.min_score is not None:
        out_df = out_df[out_df["우선순위_종합점수"] >= args.min_score]
    if args.top:
        out_df = out_df.head(args.top)

    out_df.to_excel(args.output, index=False)
    total_emails = out_df["발송대상이메일_개수"].sum()
    log(f"발송 대상 {len(out_df)}개 기관 (이메일 총 {total_emails}건) -> {args.output}")
    log("이 목록으로 메일을 보낸 뒤, mark-sent 명령으로 발송완료 기록을 남겨주세요.")


# =========================================================
# 명령: mark-sent - 발송완료 기록 (대시보드 컬럼 + 발송이력 CSV 둘 다)
# =========================================================
def cmd_mark_sent(args):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    branch_names = []
    history_rows = []

    if args.from_file:
        df = pd.read_excel(args.from_file)
        for _, row in df.iterrows():
            branch = row.get("지점명")
            org = row.get("기관명", "")
            emails_str = row.get("발송대상이메일", "")
            if isinstance(branch, str) and branch:
                branch_names.append(branch)
            if isinstance(emails_str, str):
                for e in emails_str.split(","):
                    e = e.strip()
                    if e:
                        history_rows.append(
                            {"기관명": org, "지점명": branch, "이메일": e, "발송일자": now, "캠페인명": args.campaign}
                        )
    elif args.branches:
        branch_names = [b.strip() for b in args.branches.split(",") if b.strip()]
        for b in branch_names:
            history_rows.append(
                {"기관명": "", "지점명": b, "이메일": "", "발송일자": now, "캠페인명": args.campaign}
            )
    else:
        log("--from 또는 --branches 중 하나는 반드시 지정해야 합니다.")
        return

    if not branch_names:
        log("기록할 기관이 없습니다.")
        return

    write_send_status(branch_names, args.campaign, args.dashboard)

    history = load_history(args.history)
    new_df = pd.DataFrame(history_rows)
    combined = pd.concat([history, new_df], ignore_index=True)
    combined = combined.drop_duplicates(subset=["지점명", "이메일", "캠페인명"], keep="first")
    save_history(combined, args.history)
    log(f"발송이력 CSV에도 {len(new_df)}건 추가 기록 (누적 총 {len(combined)}건) -> {args.history}")


# =========================================================
# 명령: status - 현황 요약
# =========================================================
def cmd_status(args):
    if os.path.exists(args.dashboard):
        df = load_rank_sheet(args.dashboard)
        contactable = df[df["담당자_이메일_통합모음"].fillna("").str.strip() != ""]
        sent_df = contactable[contactable["발송여부"].apply(is_sent)]
        remaining = len(contactable) - len(sent_df)
        print(f"연락 가능한 기관: {len(contactable)}개")
        print(f"  발송 완료(Y 체크됨): {len(sent_df)}개")
        print(f"  아직 발송 안 함: {remaining}개")

    history = load_history(args.history)
    print(f"\n발송이력 CSV 총 {len(history)}건 (고유 기관 {history['지점명'].nunique() if len(history) else 0}개)")
    if len(history):
        print(history["캠페인명"].value_counts().to_string())
        print("\n최근 발송 5건:")
        print(history.sort_values("발송일자", ascending=False).head(5).to_string(index=False))


def main():
    parser = argparse.ArgumentParser(description="아웃바운드 메일 발송이력 관리")
    parser.add_argument("--dashboard", default=DASHBOARD_PATH, help="대시보드 xlsx 경로")
    parser.add_argument("--history", default=HISTORY_PATH, help="발송이력 csv 경로")
    sub = parser.add_subparsers(dest="command", required=True)

    p_next = sub.add_parser("next", help="다음 발송 대상 뽑기")
    p_next.add_argument("--top", type=int, default=None, help="상위 N개 기관만")
    p_next.add_argument("--min-score", type=float, default=None, help="우선순위 종합점수 최소값")
    p_next.add_argument("--output", default=os.path.join(OUT_DIR, "발송대상_다음배치.xlsx"))
    p_next.set_defaults(func=cmd_next)

    p_mark = sub.add_parser("mark-sent", help="발송완료 기록 (대시보드 발송여부 컬럼 + 발송이력 CSV)")
    p_mark.add_argument("--from", dest="from_file", default=None, help="next로 만든 xlsx 파일 경로")
    p_mark.add_argument("--branches", default=None, help="콤마로 구분한 지점명 목록 (수동 지정 시)")
    p_mark.add_argument("--campaign", default="", help="캠페인명 (예: '2026-08 1차 제안')")
    p_mark.set_defaults(func=cmd_mark_sent)

    p_status = sub.add_parser("status", help="발송 현황 요약")
    p_status.set_defaults(func=cmd_status)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()

