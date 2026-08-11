#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_dashboard.py
=========================================================
'정부지원사업_대시보드.xlsx'를 매번 처음부터 다시 만드는 스크립트.

입력 (3개, data/ 폴더에 위치):
  1. institutions_master.csv  - ① 26년 기준 공공기관 현황 (고정, 5,974행)
  2. yearly_history.xlsx      - ② 기업마당 1년치 (고정, 1,570행)
  3. daily_accumulated.xlsx   - ③ 매일 누적 데이터 (매일 커짐)

출력 (outputs/ 폴더):
  정부지원사업_대시보드.xlsx (5개 시트: 원본 3개 + 파생 2개)

사용법:
  python3 build_dashboard.py

기관 매칭 로직 (핵심):
  ②/③의 '수행기관명'을 ①의 '지점명'과 매칭합니다. 지점명은 5,974건 중
  5,972건이 유니크해서 사실상 기본키로 쓸 수 있습니다 (기관명은 한 기관이
  여러 지점을 가질 수 있어 중복이 많음 - 예: '서울특별시상수도' 계열).

  단순 완전일치만 하면 매칭률이 낮습니다 (② 30%, ③ 60%) - 실제로 많은
  경우 ①에는 '재단법인 OOO', '사단법인 OOO' 처럼 법인격 접두어가 붙어있고
  ②/③에는 접두어 없이 'OOO'만 있기 때문입니다. 이 접두어를 제거하고
  비교하면 매칭률이 크게 올라갑니다 (③ 기준 60% -> 75%).

  그래도 매칭이 안 되는 이름들은 대부분:
  - 실제로 공공기관이 아닌 민간단체/협회 (대한상공회의소, 한국산업기술진흥협회 등)
  - '직접수행', '기초자치단체' 같은 기관명이 아닌 플레이스홀더
  이런 경우이며, ①(공공기관 현황) 자체가 커버하는 범위 밖이라 정상입니다.

점수 계산 (명확하고 설명 가능하게 새로 정의 - 원본 대시보드의 불투명한
점수 공식을 그대로 재현하지 않고, 아래처럼 재정의했습니다):

  사업규모점수 = 그 기관이 매일데이터(③)에 등록한 모든 공고의
                '지원규모' 텍스트에서 파싱한 금액(원 단위, 최댓값 기준)을
                백만원 단위로 환산해 합산한 값.
                (예: "50백만원"과 "8개사 내외" 언급 → 50만 반영,
                 인원수는 점수에 넣지 않음 - 금액 비교가 아니라서 섞으면
                 오히려 불명확해지기 때문)

  우선순위 종합점수 = 아래 3개 지표를 각각 0~100 백분위로 정규화한 뒤
                     가중평균 (가중치는 스크립트 상단 상수로 명시):
    - 서류적합도 (원클릭 매칭 서류 합계) 40%
    - 사업규모점수 35%
    - 활동성 (매일데이터 등록 공고수) 25%

  담당 이메일이 하나도 없는 기관은 우선순위와 별개로 '연락가능여부' 컬럼에
  표시해서, 점수는 높은데 정작 연락할 방법이 없는 경우를 바로 알 수 있게
  했습니다.
=========================================================
"""

import os
import re
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# 0. 경로 / 가중치 설정
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(os.path.dirname(BASE_DIR), "data")
OUT_DIR = os.path.join(os.path.dirname(BASE_DIR), "outputs")
os.makedirs(OUT_DIR, exist_ok=True)

PATH_INSTITUTIONS = os.path.join(DATA_DIR, "institutions_master.csv")
PATH_YEARLY = os.path.join(DATA_DIR, "yearly_history.xlsx")
PATH_DAILY = os.path.join(DATA_DIR, "daily_accumulated.xlsx")
PATH_OUTPUT = os.path.join(OUT_DIR, "정부지원사업_대시보드.xlsx")
RANK_SHEET_NAME = "대시보드_우선순위_랭킹"

# 우선순위 종합점수 가중치 (합이 1.0이 되도록. 필요시 이 값만 조정하면 됨)
WEIGHT_DOCS = 0.40      # 서류적합도 (원클릭 매칭 서류 합계)
WEIGHT_BUDGET = 0.35    # 사업규모점수
WEIGHT_ACTIVITY = 0.25  # 활동성 (매일데이터 등록 공고수)

# 법인격 접두어 (매칭 시 제거하고 비교)
ENTITY_PREFIXES = ["재단법인 ", "사단법인 ", "특수법인 ", "주식회사 ", "(재)", "(사)", "(주)"]

# 금액 단위 변환 (우선 공식 정의 - README에 그대로 문서화됨)
_WON_UNIT_MULTIPLIER = {
    "억": 100_000_000, "천만": 10_000_000, "백만": 1_000_000,
    "천": 1_000, "만": 10_000,
}
_WON_COMPOUND_RE = re.compile(r"(?<![\d.])(\d+)\s*억\s*(?:(\d+)\s*(천만|백만|천|만))?\s*원")
_WON_AMOUNT_RE = re.compile(r"(\d[\d,]*(?:\.\d+)?)\s*(억|천만|백만|천|만)?\s*원")

# 원클릭 매칭 서류 컬럼 목록 (매일데이터 xlsx의 Y/N 컬럼과 동일해야 함)
DOC_FLAG_COLUMNS = [
    "사업자등록증명", "표준재무제표증명", "부가가치세과세표준증명", "납부내역증명", "납세증명",
    "면세사업자수입금액증명원", "사업자등록증명(영문)", "4대보험완납증명서", "나의부동산정보",
    "지방세 납세증명", "지방세 세목별 과세증명", "법인 등기부등본", "법인세 신고내역",
    "종합소득세 신고내역", "부가세 신고내역", "원천세 신고내역", "거래처별합계표",
    "산업재해율확인서", "산업재해 요양승인 및 반려여부 확인서",
]


def log(msg):
    print(f"[build_dashboard] {msg}", flush=True)


# =========================================================
# 1. 기관명 정규화 / 매칭
# =========================================================
def normalize_org_name(name):
    if not isinstance(name, str):
        return ""
    n = name.strip()
    for p in ENTITY_PREFIXES:
        n = n.replace(p, "")
    return re.sub(r"\s+", "", n)  # 공백도 제거해서 비교 (띄어쓰기 차이 방지)


def parse_won_amount(text):
    """지원규모 텍스트에서 가장 큰 금액을 원(KRW) 단위로 파싱 (근사치)."""
    if not isinstance(text, str) or not text:
        return 0
    best = 0.0
    for m in _WON_COMPOUND_RE.finditer(text):
        eok = int(m.group(1))
        value = eok * 100_000_000
        if m.group(2) and m.group(3):
            value += int(m.group(2)) * _WON_UNIT_MULTIPLIER.get(m.group(3), 1)
        best = max(best, value)
    for m in _WON_AMOUNT_RE.finditer(text):
        try:
            num = float(m.group(1).replace(",", ""))
        except ValueError:
            continue
        best = max(best, num * _WON_UNIT_MULTIPLIER.get(m.group(2), 1))
    return int(round(best))


def percentile_rank(series):
    """0~100 백분위 (동점은 평균 순위). 값이 전부 0/동일하면 전부 0 반환."""
    if series.max() == series.min():
        return pd.Series(0.0, index=series.index)
    return series.rank(pct=True) * 100


# =========================================================
# 2. 데이터 로드
# =========================================================
def load_sources():
    log("원본 데이터 로딩 중...")
    df_inst = pd.read_csv(PATH_INSTITUTIONS, encoding="utf-8-sig")
    df_yearly = pd.read_excel(PATH_YEARLY)
    df_daily = pd.read_excel(PATH_DAILY)
    log(f"  ① 공공기관현황: {len(df_inst)}행")
    log(f"  ② 1년치 지원사업: {len(df_yearly)}행")
    log(f"  ③ 매일 누적데이터: {len(df_daily)}행")
    return df_inst, df_yearly, df_daily


# =========================================================
# 3. 매칭 인덱스 구성
# =========================================================
def build_institution_index(df_inst):
    """정규화된 지점명 -> institutions_master 행 인덱스(첫 매칭 우선)."""
    idx_map = {}
    for i, row in df_inst.iterrows():
        key = normalize_org_name(row.get("지점명"))
        if key and key not in idx_map:
            idx_map[key] = i
    return idx_map


def match_name(name, idx_map):
    key = normalize_org_name(name)
    return idx_map.get(key) if key else None


# =========================================================
# 4. 기관별 집계
# =========================================================
def aggregate(df_inst, df_yearly, df_daily, idx_map):
    log("기관별 집계 중...")

    n = len(df_inst)
    yearly_count = np.zeros(n, dtype=int)
    yearly_emails = [set() for _ in range(n)]

    daily_count = np.zeros(n, dtype=int)
    daily_doc_sum = np.zeros(n, dtype=int)
    daily_budget_sum = np.zeros(n, dtype=float)  # 백만원 단위
    daily_emails = [set() for _ in range(n)]
    daily_phones = [set() for _ in range(n)]

    unmatched_yearly = set()
    unmatched_daily = set()

    for _, row in df_yearly.iterrows():
        i = match_name(row.get("사업수행기관명"), idx_map)
        if i is None:
            if isinstance(row.get("사업수행기관명"), str):
                unmatched_yearly.add(row["사업수행기관명"])
            continue
        yearly_count[i] += 1
        for col in ["문의처(이메일주소)", "사업신청 방법(이메일주소)"]:
            val = row.get(col)
            if isinstance(val, str) and "@" in val:
                for e in re.split(r"[,\s;]+", val):
                    e = e.strip()
                    if "@" in e and "없음" not in e:
                        yearly_emails[i].add(e)

    for _, row in df_daily.iterrows():
        i = match_name(row.get("수행기관명(excInsttNm)"), idx_map)
        if i is None:
            if isinstance(row.get("수행기관명(excInsttNm)"), str):
                unmatched_daily.add(row["수행기관명(excInsttNm)"])
            continue
        daily_count[i] += 1
        daily_doc_sum[i] += sum(1 for c in DOC_FLAG_COLUMNS if row.get(c) == "Y")
        daily_budget_sum[i] += parse_won_amount(row.get("지원규모")) / 1_000_000

        email = row.get("담당이메일")
        if isinstance(email, str) and "@" in email:
            daily_emails[i].add(email.strip())
        phone = row.get("담당전화번호")
        if isinstance(phone, str) and phone.strip():
            daily_phones[i].add(phone.strip())

    log(f"  ② 매칭 안 된 기관명 {len(unmatched_yearly)}종 (공공기관 목록 밖 - 정상)")
    log(f"  ③ 매칭 안 된 기관명 {len(unmatched_daily)}종 (공공기관 목록 밖 - 정상)")

    result = df_inst.copy()
    result["1년치_지원사업_공고수"] = yearly_count
    result["매일업로드_공고수"] = daily_count
    result["총_필요서류_합계"] = daily_doc_sum
    result["사업규모점수"] = daily_budget_sum.round(1)

    combined_emails = []
    for i in range(n):
        combined = sorted(yearly_emails[i] | daily_emails[i])
        combined_emails.append(", ".join(combined))
    result["담당자_이메일_통합모음"] = combined_emails
    result["담당자_이메일_개수"] = [len(yearly_emails[i] | daily_emails[i]) for i in range(n)]

    return result, unmatched_yearly, unmatched_daily


# =========================================================
# 5. 우선순위 점수 계산
# =========================================================
def load_existing_send_status(path=PATH_OUTPUT):
    """
    이전에 생성된 대시보드 파일에서 '발송여부'/'발송일자' 값을 읽어옵니다.
    build_dashboard.py는 매번 대시보드를 처음부터 다시 만들기 때문에, 이걸
    안 하면 엑셀에 직접 체크해둔 발송여부가 재생성할 때마다 사라집니다.
    반환값: {지점명: {"발송여부": ..., "발송일자": ...}}
    """
    if not os.path.exists(path):
        return {}
    try:
        old_df = pd.read_excel(path, sheet_name=RANK_SHEET_NAME, header=3)
    except Exception as e:
        log(f"기존 대시보드에서 발송상태를 읽지 못했습니다(최초 생성이면 정상): {e}")
        return {}

    status = {}
    for _, row in old_df.iterrows():
        branch = row.get("지점명")
        if not isinstance(branch, str) or not branch:
            continue
        status[branch] = {
            "발송여부": row.get("발송여부", "") or "",
            "발송일자": row.get("발송일자", "") or "",
        }
    log(f"기존 대시보드에서 발송상태 {len(status)}건 불러옴 (재생성해도 유지됩니다)")
    return status


def apply_send_status(df, status_map):
    """기존 발송여부/발송일자를 지점명 기준으로 새 데이터프레임에 병합."""
    df["발송여부"] = df["지점명"].map(lambda b: status_map.get(b, {}).get("발송여부", ""))
    df["발송일자"] = df["지점명"].map(lambda b: status_map.get(b, {}).get("발송일자", ""))
    return df


def compute_priority(df):
    log("우선순위 점수 계산 중...")
    docs_pct = percentile_rank(df["총_필요서류_합계"])
    budget_pct = percentile_rank(df["사업규모점수"])
    activity_pct = percentile_rank(df["매일업로드_공고수"])

    df["우선순위_종합점수"] = (
        WEIGHT_DOCS * docs_pct + WEIGHT_BUDGET * budget_pct + WEIGHT_ACTIVITY * activity_pct
    ).round(1)

    df["연락가능여부"] = np.where(
        df["담당자_이메일_개수"] > 0, "가능",
        np.where(df["이메일"].fillna("-") != "-", "대표이메일만", "연락처없음")
    )

    df = df.sort_values("우선순위_종합점수", ascending=False).reset_index(drop=True)
    df.insert(0, "제안순위", df.index + 1)
    return df


# =========================================================
# 6. 엑셀 출력
# =========================================================
HEADER_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def sanitize_for_excel(value):
    """PDF/HWP/OCR 추출 결과에 섞여 들어올 수 있는 제어문자를 제거합니다
    (bizinfo_bot.py의 동일 함수와 같은 이유 - openpyxl이 이런 문자가 있으면
    IllegalCharacterError를 던져서 저장 전체가 실패합니다)."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _write_df_sheet(wb, sheet_name, title, subtitle, df, col_widths=None, dropdown_columns=None):
    """
    dropdown_columns: {"컬럼명": ["Y", "N"]} 형태로 지정하면, 그 컬럼 전체
    (데이터 영역)에 드롭다운 목록(데이터 유효성 검사)을 넣어서 체크박스처럼
    클릭으로 값을 고를 수 있게 합니다.
    """
    ws = wb.create_sheet(sheet_name)
    ws.append([title])
    ws.append([subtitle])
    ws.append([])
    ws.append(list(df.columns))

    header_row = 4
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in df.itertuples(index=False):
        ws.append([sanitize_for_excel(v) for v in row])

    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    if dropdown_columns:
        from openpyxl.worksheet.datavalidation import DataValidation
        col_names = list(df.columns)
        last_data_row = header_row + len(df)
        for col_name, options in dropdown_columns.items():
            if col_name not in col_names:
                continue
            col_idx = col_names.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="list", formula1=f'"{",".join(options)}"', allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}{header_row + 1}:{col_letter}{last_data_row}")

    return ws


def write_dashboard(df_rank, df_inst_raw, df_yearly_raw, df_daily_raw):
    log("엑셀 파일 작성 중...")
    wb = Workbook()
    wb.remove(wb.active)

    # --- 대시보드_우선순위_랭킹 ---
    rank_cols = [
        "제안순위", "기관명", "지점명", "상위기관_주무부처_지자체", "기관유형",
        "1년치_지원사업_공고수", "매일업로드_공고수", "총_필요서류_합계",
        "사업규모점수", "우선순위_종합점수", "연락가능여부", "발송여부", "발송일자",
        "담당자_이메일_통합모음", "대표전화", "이메일",
    ]
    _write_df_sheet(
        wb, "대시보드_우선순위_랭킹",
        "B2G 솔루션 제안 우선순위 랭킹 대시보드 (전국 공공기관 현황 전수 반영)",
        "기준: 서류적합도 40% + 사업규모 35% + 활동성 25% 가중 백분위. "
        "'발송여부' 컬럼에 Y를 직접 입력하면 다음 재생성 때도 유지되고, "
        "outbound_email_manager.py의 next 명령에서 자동으로 제외됩니다.",
        df_rank[rank_cols],
        col_widths=[8, 22, 22, 20, 18, 14, 14, 12, 12, 14, 12, 10, 14, 60, 16, 24],
        dropdown_columns={"발송여부": ["Y", "N"]},
    )

    # --- 공공기관별_컨택디렉토리 ---
    contact_cols = [
        "제안순위", "기관명", "지점명", "상위기관_주무부처_지자체",
        "1년치_지원사업_공고수", "매일업로드_공고수",
        "담당자_이메일_통합모음", "대표전화", "이메일",
    ]
    _write_df_sheet(
        wb, "공공기관별_컨택디렉토리",
        "공공기관별 담당자 이메일 & 컨택 포인트 디렉토리 (전국 전수)",
        "우선순위_랭킹과 동일 순서 (제안 활동은 이 시트를 outbound 관리 스크립트와 함께 사용)",
        df_rank[contact_cols],
        col_widths=[8, 22, 22, 20, 14, 14, 60, 16, 24],
    )

    # --- 원본 3개 시트 (그대로 보존) ---
    _write_df_sheet(
        wb, "매일업로드_원천데이터",
        "매일 업로드 원천 데이터 (③ 시트 - 매일 업데이트 영역)",
        f"누적 {len(df_daily_raw)}건 (data/daily_accumulated.xlsx 기준)",
        df_daily_raw,
    )
    _write_df_sheet(
        wb, "1년치지원사업_원본",
        "기업마당 API 지원사업 1년치 업로드 리스트업 (② 시트 원본)",
        f"{len(df_yearly_raw)}건",
        df_yearly_raw,
    )
    _write_df_sheet(
        wb, "공공기관현황_원본",
        "26년 기준 공공기관 현황 마스터 (① 시트 원본 전수)",
        f"{len(df_inst_raw)}건",
        df_inst_raw,
    )

    wb.save(PATH_OUTPUT)
    log(f"저장 완료: {PATH_OUTPUT}")


def main():
    df_inst, df_yearly, df_daily = load_sources()
    idx_map = build_institution_index(df_inst)
    agg, unmatched_yearly, unmatched_daily = aggregate(df_inst, df_yearly, df_daily, idx_map)
    ranked = compute_priority(agg)

    # 기존 대시보드에 수동으로 체크해둔 발송여부/발송일자를 불러와서 유지
    send_status = load_existing_send_status()
    ranked = apply_send_status(ranked, send_status)

    write_dashboard(ranked, df_inst, df_yearly, df_daily)

    log("=== 완료 ===")
    log(f"전체 기관 수: {len(ranked)}")
    log(f"매일데이터 매칭된 기관 수(공고수>0): {(ranked['매일업로드_공고수'] > 0).sum()}")
    log(f"연락 가능(이메일 확보) 기관 수: {(ranked['연락가능여부'] == '가능').sum()}")
    top5 = ranked.head(5)[["제안순위", "지점명", "우선순위_종합점수", "매일업로드_공고수", "총_필요서류_합계", "사업규모점수"]]
    log("상위 5개 기관:\n" + top5.to_string(index=False))


if __name__ == "__main__":
    main()
